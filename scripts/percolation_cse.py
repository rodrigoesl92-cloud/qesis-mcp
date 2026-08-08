"""Percolation on the submarine-cable graph.

Closes the open item recorded at ops/CONNECTIVITY_BATCH_AUDIT_2026-07-31.md:106:
"Run percolation on the 1,489-edge graph. Remove articulation points in
betweenness order and measure component fragmentation. That produces the
propagation rule and the tipping point the Haldane framing needs."

The graph is bipartite: landing cities on one side, cables on the other, an edge
where a cable lands at a city. That is the structure the 2026-07-31 audit
reproduced independently in networkx (1,255 nodes, 1,489 edges, 76 components,
88 articulation cities matched 88 of 88).

Everything here is pure stdlib apart from openpyxl for the workbook read. numpy
is available but not used: 1,255 nodes does not need it and the machine runs
under roughly 450 MB free.

Attack orders compared:
  articulation_betweenness  the ordering the audit asked for
  betweenness_static        all cities by betweenness, computed once
  betweenness_recalc        betweenness recomputed after every removal
  degree                    cable count, the cheap adversary
  random                    averaged over seeds, the null model

Reported per step: largest component share, component count, mean component
size, and the share of city pairs still connected. The tipping point is the
removal count at which the largest component share first falls below 0.5.
"""

import json
import os
import random
from collections import defaultdict, deque

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(REPO, "data", "axes", "cse_percolation.json")
RANDOM_SEEDS = [11, 23, 37, 51, 67]


# --------------------------------------------------------------------------
# graph construction
# --------------------------------------------------------------------------

def load_graph(workbook):
    import openpyxl
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)

    adj = defaultdict(set)
    city_meta = {}

    ws = wb["cities"]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    for cid, name, country, lon, lat in rows:
        if cid is None:
            continue
        city_meta[f"c{cid}"] = {"name": name, "country": country}

    ws = wb["city_cable"]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    n_edges = 0
    for cid, name, country, cable_id in rows:
        if cid is None or cable_id is None:
            continue
        u, v = f"c{cid}", f"k{cable_id}"
        if v not in adj[u]:
            n_edges += 1
        adj[u].add(v)
        adj[v].add(u)

    published = {}
    ws = wb["chokepoints"]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    for cid, betw, is_art, name, country, lon, lat, cable_count in rows:
        if cid is None:
            continue
        published[f"c{cid}"] = {
            "betweenness": float(betw) if betw is not None else 0.0,
            "is_articulation": str(is_art).strip().lower() == "true",
        }
    wb.close()
    return adj, city_meta, published, n_edges


# --------------------------------------------------------------------------
# structure
# --------------------------------------------------------------------------

def components(adj, removed=frozenset()):
    """Connected components as a list of node sets."""
    seen = set(removed)
    out = []
    for s in adj:
        if s in seen:
            continue
        comp = set()
        dq = deque([s])
        seen.add(s)
        while dq:
            u = dq.popleft()
            comp.add(u)
            for w in adj[u]:
                if w not in seen:
                    seen.add(w)
                    dq.append(w)
        out.append(comp)
    return out


def brandes_betweenness(adj, alive):
    """Unweighted betweenness (Brandes 2001) over the induced subgraph `alive`."""
    cb = dict.fromkeys(alive, 0.0)
    for s in alive:
        stack = []
        preds = {w: [] for w in alive}
        sigma = dict.fromkeys(alive, 0.0)
        dist = dict.fromkeys(alive, -1)
        sigma[s] = 1.0
        dist[s] = 0
        dq = deque([s])
        while dq:
            v = dq.popleft()
            stack.append(v)
            for w in adj[v]:
                if w not in alive:
                    continue
                if dist[w] < 0:
                    dist[w] = dist[v] + 1
                    dq.append(w)
                if dist[w] == dist[v] + 1:
                    sigma[w] += sigma[v]
                    preds[w].append(v)
        delta = dict.fromkeys(alive, 0.0)
        while stack:
            w = stack.pop()
            for v in preds[w]:
                delta[v] += (sigma[v] / sigma[w]) * (1.0 + delta[w])
            if w != s:
                cb[w] += delta[w]
    n = len(alive)
    if n > 2:
        scale = 1.0 / ((n - 1) * (n - 2))
        for k in cb:
            cb[k] *= scale
    return cb


def articulation_points(adj, alive):
    """Hopcroft-Tarjan cut vertices, iterative to survive deep graphs."""
    disc, low, parent = {}, {}, {}
    arts = set()
    timer = [0]
    for root in alive:
        if root in disc:
            continue
        parent[root] = None
        stack = [(root, iter(adj[root]))]
        disc[root] = low[root] = timer[0]
        timer[0] += 1
        root_children = 0
        while stack:
            node, it = stack[-1]
            advanced = False
            for w in it:
                if w not in alive:
                    continue
                if w not in disc:
                    parent[w] = node
                    disc[w] = low[w] = timer[0]
                    timer[0] += 1
                    stack.append((w, iter(adj[w])))
                    if node == root:
                        root_children += 1
                    advanced = True
                    break
                if w != parent.get(node):
                    low[node] = min(low[node], disc[w])
            if not advanced:
                stack.pop()
                if stack:
                    p = stack[-1][0]
                    low[p] = min(low[p], low[node])
                    if parent.get(p) is not None and low[node] >= disc[p]:
                        arts.add(p)
        if root_children > 1:
            arts.add(root)
    return arts


def measure(adj, removed, n_cities_total):
    """Fragmentation statistics for the graph minus `removed`."""
    comps = components(adj, removed)
    city_comps = [sum(1 for x in c if x.startswith("c")) for c in comps]
    city_comps = [n for n in city_comps if n > 0]
    if not city_comps:
        return {"lcc_cities": 0, "lcc_share": 0.0, "n_components": 0,
                "mean_component": 0.0, "pair_connectivity": 0.0}
    lcc = max(city_comps)
    alive_cities = sum(city_comps)
    # Share of surviving city pairs that remain mutually reachable.
    pairs_total = alive_cities * (alive_cities - 1) / 2 if alive_cities > 1 else 0
    pairs_conn = sum(n * (n - 1) / 2 for n in city_comps)
    return {
        "lcc_cities": lcc,
        "lcc_share": round(lcc / n_cities_total, 4),
        "n_components": len(city_comps),
        "mean_component": round(alive_cities / len(city_comps), 2),
        "pair_connectivity": round(pairs_conn / pairs_total, 4) if pairs_total else 0.0,
    }


# --------------------------------------------------------------------------
# attack campaigns
# --------------------------------------------------------------------------

def run_campaign(adj, order, n_cities_total, max_steps, label, recalc=False,
                 cities=None):
    """Remove nodes one at a time, recording fragmentation after each."""
    removed = set()
    curve = [dict(step=0, removed=None, **measure(adj, removed, n_cities_total))]
    if recalc:
        alive = set(cities)
        for step in range(1, max_steps + 1):
            live_nodes = {n for n in adj if n not in removed}
            cb = brandes_betweenness(adj, live_nodes)
            cand = [(cb.get(c, 0.0), c) for c in alive if c not in removed]
            if not cand:
                break
            cand.sort(reverse=True)
            victim = cand[0][1]
            removed.add(victim)
            curve.append(dict(step=step, removed=victim,
                              **measure(adj, removed, n_cities_total)))
    else:
        for step, victim in enumerate(order[:max_steps], start=1):
            removed.add(victim)
            curve.append(dict(step=step, removed=victim,
                              **measure(adj, removed, n_cities_total)))
    return {"label": label, "curve": curve}


def tipping_point(curve, threshold=0.5):
    """First removal count where largest-component share drops below threshold."""
    base = curve[0]["lcc_share"]
    for row in curve:
        if row["lcc_share"] < threshold * base:
            return {"steps": row["step"], "lcc_share": row["lcc_share"],
                    "relative_to_baseline": round(row["lcc_share"] / base, 4)}
    return None


def main():
    scratch = os.environ.get("QESIS_SCRATCH", "")
    workbook = os.path.join(scratch, "connectivity_L1_full.xlsx")
    adj, city_meta, published, n_edges = load_graph(workbook)

    cities = [n for n in adj if n.startswith("c")]
    cables = [n for n in adj if n.startswith("k")]
    n_cities_total = len(cities)
    alive = set(adj)

    print(f"nodes {len(adj)} (cities {n_cities_total}, cables {len(cables)})  "
          f"edges {n_edges}")

    # Independent recomputation before anything is trusted.
    base_comps = components(adj)
    arts = articulation_points(adj, alive)
    art_cities = {a for a in arts if a.startswith("c")}
    pub_arts = {k for k, v in published.items() if v["is_articulation"]}
    print(f"components {len(base_comps)}  articulation cities recomputed "
          f"{len(art_cities)}  published {len(pub_arts)}  "
          f"overlap {len(art_cities & pub_arts)}")

    cb = brandes_betweenness(adj, alive)

    # Orderings.
    art_order = sorted(art_cities, key=lambda c: -published.get(c, {}).get("betweenness", 0.0))
    betw_order = sorted(cities, key=lambda c: -cb.get(c, 0.0))
    deg_order = sorted(cities, key=lambda c: -len(adj[c]))

    max_steps = len(art_order)  # 88 articulation cities
    campaigns = []

    campaigns.append(run_campaign(adj, art_order, n_cities_total, max_steps,
                                  "articulation_betweenness"))
    campaigns.append(run_campaign(adj, betw_order, n_cities_total, max_steps,
                                  "betweenness_static"))
    campaigns.append(run_campaign(adj, deg_order, n_cities_total, max_steps,
                                  "degree"))
    campaigns.append(run_campaign(adj, None, n_cities_total, max_steps,
                                  "betweenness_recalc", recalc=True, cities=cities))

    # Random null, averaged over seeds.
    rand_curves = []
    for seed in RANDOM_SEEDS:
        rng = random.Random(seed)
        order = cities[:]
        rng.shuffle(order)
        rand_curves.append(run_campaign(adj, order, n_cities_total, max_steps,
                                        f"random_{seed}")["curve"])
    avg = []
    for i in range(max_steps + 1):
        avg.append({
            "step": i,
            "lcc_share": round(sum(c[i]["lcc_share"] for c in rand_curves) / len(rand_curves), 4),
            "n_components": round(sum(c[i]["n_components"] for c in rand_curves) / len(rand_curves), 2),
            "pair_connectivity": round(sum(c[i]["pair_connectivity"] for c in rand_curves) / len(rand_curves), 4),
        })
    campaigns.append({"label": "random_mean", "curve": avg})

    result = {
        "generated_by": "scripts/percolation_cse.py",
        "closes": "ops/CONNECTIVITY_BATCH_AUDIT_2026-07-31.md:106",
        "graph": {
            "nodes": len(adj), "cities": n_cities_total, "cables": len(cables),
            "edges": n_edges, "components": len(base_comps),
            "articulation_cities_recomputed": len(art_cities),
            "articulation_cities_published": len(pub_arts),
            "articulation_overlap": len(art_cities & pub_arts),
        },
        "campaigns": campaigns,
        "tipping_points": {
            c["label"]: tipping_point(c["curve"]) for c in campaigns
        },
        "top_chokepoints": [
            {"node": c, "city": city_meta.get(c, {}).get("name"),
             "country": city_meta.get(c, {}).get("country"),
             "betweenness": round(cb.get(c, 0.0), 6),
             "cables": len(adj[c]), "is_articulation": c in art_cities}
            for c in betw_order[:15]
        ],
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    json.dump(result, open(OUT, "w", encoding="utf-8"), indent=2)
    print(f"wrote {OUT}")
    return result


if __name__ == "__main__":
    main()
